"""Экспорт данных в Excel (.xlsx) для админ-панели."""
import io
import os

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


def _pdf_font():
    candidates = [
        os.path.join(os.environ.get('WINDIR', ''), 'Fonts', 'arial.ttf'),
        '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
    ]
    for path in candidates:
        if os.path.exists(path):
            try:
                pdfmetrics.registerFont(TTFont('CourseUnicode', path))
                return 'CourseUnicode'
            except (OSError, ValueError):
                pass
    return 'Helvetica'


def _autosize(ws):
    for col in ws.columns:
        values = [str(c.value) for c in col if c.value is not None]
        length = max((len(v) for v in values), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max(length + 2, 12), 40)


def export_student_progress_xlsx(courses_data):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Прогресс'
    headers = ['Тренинг', 'Модуль', 'Уроков просмотрено', 'Всего уроков', 'Есть тест', 'Балл теста', 'Статус']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for course in courses_data:
        for m in course['modules']:
            if m['completed']:
                mstatus = 'Пройден'
            elif not m['unlocked']:
                mstatus = 'Заблокирован'
            else:
                mstatus = 'В процессе'
            ws.append([
                course['course_title'],
                m['title'],
                m['lessons_watched'],
                m['lessons_total'],
                'Да' if m['has_test'] else 'Нет',
                m['test_best_score'] if m['test_best_score'] is not None else '—',
                mstatus,
            ])

    _autosize(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_students_xlsx(students):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Ученики'
    headers = ['Имя', 'Фамилия', 'Email', 'Телефон', 'Статус', 'Дата регистрации', 'Курсы']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for s in students:
        ws.append([
            s.first_name,
            s.last_name,
            s.email,
            s.phone or '—',
            'Активен' if s.is_active_student else 'Не активен',
            s.date_joined.strftime('%d.%m.%Y'),
            ', '.join(e.course.title for e in s.enrollments.all()),
        ])

    _autosize(ws)
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def export_student_progress_pdf(student, courses_data):
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    _, height = A4
    y = height - 48
    font_name = _pdf_font()

    def write(text, size=10, bold=False):
        nonlocal y
        if y < 48:
            pdf.showPage()
            y = height - 48
        pdf.setFont(font_name, size)
        pdf.drawString(42, y, str(text)[:120])
        y -= size + 6

    write('Отчёт о прогрессе ученика', 16, True)
    write(f'{student.get_full_name() or student.email} · {student.email}')
    y -= 8
    for course in courses_data:
        write(course['course_title'], 13, True)
        for module in course['modules']:
            state = 'пройден' if module['completed'] else 'не завершён'
            write(f"  {module['title']}: {module['lessons_watched']}/{module['lessons_total']} уроков, {state}")
            for attempt in module.get('test_attempts', []):
                write(
                    f"    Попытка теста: {attempt['score_percent']}% · "
                    f"{'сдан' if attempt['passed'] else 'не сдан'} · {attempt['submitted_at']}",
                    9,
                )
        y -= 6
    pdf.save()
    buffer.seek(0)
    return buffer


def export_students_pdf(students):
    buffer = io.BytesIO()
    pdf = canvas.Canvas(buffer, pagesize=A4)
    _, height = A4
    font_name = _pdf_font()
    y = height - 48

    def write(text, size=10, bold=False):
        nonlocal y
        if y < 48:
            pdf.showPage()
            y = height - 48
        pdf.setFont(font_name, size)
        pdf.drawString(42, y, str(text)[:120])
        y -= size + 6

    write('Список учеников', 16, True)
    y -= 8
    for index, student in enumerate(students, 1):
        courses = ', '.join(e.course.title for e in student.enrollments.all()) or 'без курса'
        write(
            f'{index}. {student.get_full_name() or "Без имени"} · {student.email} · '
            f'{"Активен" if student.is_active_student else "Не активен"}'
        )
        write(f'   Телефон: {student.phone or "—"} · Курсы: {courses}', 9)
    pdf.save()
    buffer.seek(0)
    return buffer
